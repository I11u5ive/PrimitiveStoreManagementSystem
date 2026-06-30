from openpyxl import load_workbook

class Product:
    def __init__(self, name: str, toy_type: str, price: int | float, quantity: int):
        self.name = name
        self.toy_type = toy_type
        self.price = price
        self.quantity = quantity

    # Getters:

    @property
    def price(self) -> int | float:
        return self.__price

    @property
    def quantity(self) -> int:
        return self.__quantity

    # Setters:

    @price.setter
    def price(self, new_price: int | float):
        if new_price >= 0:
            self.__price = new_price
        else:
            raise ValueError('Price cannot be less than 0.')

    @quantity.setter
    def quantity(self, new_quantity: int):
        if new_quantity >= 0:
            self.__quantity = new_quantity
        else:
            raise ValueError('Quantity cannot be less than 0.')

    #Other methods:

    def __str__(self):
        return f'{self.name} | {self.toy_type} | {self.price}$ | {self.quantity}'


class Order:

    def __init__(self):
        self.toys = []
        self.__total_price = 0

    # Getters:

    @property
    def total_price(self) -> int | float:
        return self.__total_price

    # Other methods:

    def __str__(self):
        text = 'Order:\n'

        for toy, quantity in self.toys:
            text += f"{toy} x {quantity}\n"

        text += f"Total price: {self.total_price}$"

        return text

    def __update_total_price(self):
        self.__total_price = sum(
            toy.price * quantity
            for toy, quantity in self.toys
        )

    def add_toy(self, new_toy: Product, new_quantity = 1):
        if new_quantity < 1:
            raise ValueError("Quantity cannot be less than 1.")

        if new_toy.quantity >= new_quantity:

            old_quantity = new_toy.quantity
            new_toy.quantity = new_toy.quantity - new_quantity
            self.toys.append((new_toy, new_quantity))
            self.__update_total_price()

            print(
                f"Added {new_quantity} items of {new_toy.name} "
                f"(was {old_quantity}) to order."
            )
        else:
            print(f'Not enough {new_toy.name} quantity in stock.')


class Customer:

    def __init__(self, name: str, email: str):
        self.name = name
        self.email = email
        self.__order = []

    # Getters:

    @property
    def email(self) -> str:
        return self.__email

    # Setters:

    @email.setter
    def email(self, new_email: str):
        if '@' in new_email:
            self.__email = new_email
        else:
            raise ValueError('Email must contain @.')

    # Other methods:

    def __str__(self):
        return f'{self.name} ({self.email})'

    def add_order(self, new_order: Order):
        if len(new_order.toys) > 0:
            self.__order.append(new_order)
        else:
            raise ValueError('Order must contain at least 1 element')

    def print_order(self, index):
        return str(self.__order[index])


workbook = load_workbook('store.xlsx')

# Toys:

sheet_toys = workbook['toys']

list_toys = []

for row in sheet_toys.iter_rows(min_row=2, values_only=True):
    name, toys_type, price, quantity = row
    list_toys.append(Product(name, toys_type, price, quantity))

# Customers:

sheet_customers = workbook['customers']

list_customers = []

for row in sheet_customers.iter_rows(min_row=2, values_only=True):
    name, email = row
    list_customers.append(Customer(name, email))

# Printing products:

print('Products:\n')

for number, toy in enumerate(list_toys, start=1):
    print(number, toy)

print('\nCustomers:\n')
for number, customer in enumerate(list_customers, start=1):
    print(number, customer)

# Order creating:

customer = list_customers[0]

order = Order()

order.add_toy(list_toys[0], 2)
order.add_toy(list_toys[2], 3)

customer.add_order(order)

print('\nCustomer`s order:\n')
print(customer.name)
print(customer.print_order(0))

print('\nProducts in stock:\n')

for toy in list_toys:
    print(toy)